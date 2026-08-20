"""Tests cho BroadcastReport và CSV/XLSX exporter."""
import csv

from openpyxl import load_workbook

from app.core import report_export as exporter
from app.core.broadcast_result import BroadcastReport, RecipientResult, RecipientStatus


def sample_report() -> BroadcastReport:
    report = BroadcastReport(
        avd_name="avd_1",
        serial="emulator-5554",
        started_at="2026-08-20T07:00:00+00:00",
        completed_at="2026-08-20T07:01:00+00:00",
    )
    report.add(RecipientResult(
        avd_name="avd_1",
        serial="emulator-5554",
        phone="84987654321",
        status=RecipientStatus.SUCCESS,
        attempts=1,
        elapsed_seconds=3.25,
    ))
    report.add(RecipientResult(
        avd_name="avd_1",
        serial="emulator-5554",
        phone="84900000000",
        status=RecipientStatus.FAILED,
        attempts=3,
        elapsed_seconds=8.5,
        error_code="WhatsAppError",
        error_message="=dangerous-formula",
    ))
    return report


def test_broadcast_report_summary_counts():
    report = sample_report()
    assert report.attempted_count == 2
    assert report.success_count == 1
    assert report.failed_count == 1
    assert report.partial_count == 0
    assert report.cancelled_count == 0


def test_export_csv_writes_recipient_rows_and_neutralizes_formula(tmp_path):
    report = sample_report()
    target = exporter.export_csv(report, tmp_path / "report.csv")

    with open(target, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 2
    assert rows[0]["phone"] == "84987654321"
    assert rows[1]["status"] == "failed"
    assert rows[1]["error_message"] == "'=dangerous-formula"


def test_export_xlsx_contains_summary_and_recipient_sheets(tmp_path):
    report = sample_report()
    target = exporter.export_xlsx(report, tmp_path / "report.xlsx")

    wb = load_workbook(target, data_only=False)
    assert wb.sheetnames == ["Summary", "Recipients"]
    recipients = wb["Recipients"]
    assert recipients.max_row == 3
    assert recipients.cell(row=2, column=7).value == "84987654321"
    assert recipients.cell(row=3, column=12).value == "'=dangerous-formula"


def test_preflight_failure_exports_without_recipient_rows(tmp_path):
    report = BroadcastReport(
        avd_name="avd_1",
        serial="emulator-5554",
        preflight_ok=False,
        preflight_errors=["device offline"],
    )
    report.finish()

    target = exporter.export_csv(report, tmp_path / "preflight.csv")
    with open(target, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 1
    assert rows[0]["error_code"] == "PREFLIGHT_FAILED"
    assert "device offline" in rows[0]["error_message"]


def test_default_report_path_uses_reports_directory(monkeypatch, tmp_path):
    report = sample_report()
    monkeypatch.setattr(exporter.paths, "reports_dir", lambda: tmp_path)

    path = exporter.default_report_path(report, "xlsx")

    assert path.parent == tmp_path
    assert path.suffix == ".xlsx"
    assert "avd_1" in path.name
